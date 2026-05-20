import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'docs\INV-Modbus地址表3KU（外发）V1.20.xlsx', data_only=True)

md_notes = {}
with open(r'docs\registers-map.md', encoding='utf-8') as f:
    for line in f:
        if line.startswith('|') and not line.startswith('|---') and not 'Initial Loger query' in line:
            parts = [p.strip() for p in line.split('|')]
            if len(parts) >= 11:
                dec_str = parts[5].strip() # 5 is the Dec column because cols are: 0:empty, 1:Initial, 2:Regular, 3:Marked, 4:Hex, 5:Dec, 6:Name, 7:Type, 8:Unit, 9:Perm, 10:Note, 11:empty
                try:
                    dec = int(dec_str)
                    note = parts[10].strip()
                    md_notes[dec] = note
                except ValueError:
                    pass

mismatches = 0
for sn in wb.sheetnames:
    if 'Version' in sn or 'format' in sn:
        if 'WF_Version' not in sn:
            continue
    ws = wb[sn]
    for r in ws.iter_rows(values_only=True):
        if r[1] is not None:
            try:
                dec = int(r[1])
                excel_usage = str(r[8]).strip() if len(r) > 8 and r[8] is not None else ''
                md_usage = md_notes.get(dec, '')
                # Clean clean
                excel_usage_clean = excel_usage.replace('|', '/').replace('\n', ' ').strip()
                if excel_usage_clean and not md_usage:
                    print(f'Mismatch on Dec {dec} in {sn.strip()}: Excel has "{excel_usage_clean}" but Markdown is empty!')
                    mismatches += 1
            except ValueError:
                pass

print('Total mismatches:', mismatches)
