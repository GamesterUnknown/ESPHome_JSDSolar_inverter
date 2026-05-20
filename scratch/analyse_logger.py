import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'docs\INV-Modbus地址表3KU（外发）V1.20.xlsx', data_only=True)

logger_lines = [
    '01 03 00 00 00 0B 04 0D',
    '01 03 00 10 00 30 44 1B',
    '01 03 00 40 00 5E C5 E6',
    '01 03 01 40 00 05 85 E1',
    '01 03 01 50 00 34 45 F0',
    '01 03 01 90 00 13 05 D6',
    '01 03 01 B0 00 53 05 EC',
    '01 03 02 10 00 47 05 85',
    '01 03 02 60 00 22 C4 75',
    '01 03 02 A0 00 36 C4 46',
    '01 03 03 20 00 02 C5 85',
    '01 03 03 30 00 05 85 82',
    '01 03 03 50 00 08 44 59',
    '01 03 03 60 00 06 C5 92',
    '01 03 10 00 00 0D 80 CF',
    '01 03 11 00 00 0A C0 F1',
    '01 03 12 00 00 10 41 7E',
    '01 03 20 00 00 03 0E 0B',
    '01 03 41 00 00 40 50 06',
    '01 03 41 40 00 0C 50 27'
]

requested_addrs = set()
for line in logger_lines:
    parts = line.split()
    start = int(''.join(parts[2:4]), 16)
    count = int(''.join(parts[4:6]), 16)
    for addr in range(start, start + count):
        requested_addrs.add(addr)

sheet_stats = {}
for sn in wb.sheetnames:
    if sn in (' Version information', ' Communication format'):
        continue
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = None
    for i, row in enumerate(rows):
        if row and any(str(c).lower().strip() == 'address in decimal' for c in row if c is not None):
            hdr_idx = i
            break
    if hdr_idx is None:
        continue
    
    sheet_name = sn.strip()
    sheet_stats[sheet_name] = {'total': 0, 'requested': 0, 'not_requested': 0, 'examples': []}
    
    for row in rows[hdr_idx + 1:]:
        if row[1] is None:
            continue
        try:
            addr_dec = int(row[1])
            addr_hex = str(row[2]).upper().zfill(4) if row[2] else '????'
            addr_int = int(addr_hex, 16)
        except (TypeError, ValueError):
            continue
            
        sheet_stats[sheet_name]['total'] += 1
        meaning = str(row[5]) if len(row) > 5 and row[5] is not None else 'Empty'
        
        if addr_int in requested_addrs:
            sheet_stats[sheet_name]['requested'] += 1
        else:
            sheet_stats[sheet_name]['not_requested'] += 1
            sheet_stats[sheet_name]['examples'].append((f'0x{addr_int:04X}', meaning))

print("| Назва аркуша | Всього регістрів в Excel | Запитано логером | Не запитано логером | Приклади не запитаних регістрів |")
print("|---|---|---|---|---|")
for sn in wb.sheetnames:
    sn_stripped = sn.strip()
    if sn_stripped not in sheet_stats:
        continue
    stats = sheet_stats[sn_stripped]
    ex_str = '; '.join([f'{addr} ({mean})' for addr, mean in stats['examples'][:2]])
    if len(stats['examples']) > 2:
        ex_str += ' ...'
    print(f"| {sn_stripped} | {stats['total']} | {stats['requested']} | {stats['not_requested']} | {ex_str} |")
