import openpyxl, sys, io
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(
    r'docs\INV-Modbus地址表3KU（外发）V1.20.xlsx', data_only=True
)

# ── Addresses already implemented as sensors in jsd-solar-inverter.yaml ──────
SENSOR_ADDRS = {
    # WF_Version – startup read (SN, versions, identifiers)
    0x0000, 0x0001, 0x0002, 0x0003, 0x0004,
    0x0005, 0x0006, 0x0007, 0x0008, 0x0009,  # SN code
    0x0010, 0x0011,  # Protocol version H+L
    0x0012, 0x0013,  # Power board HW H+L
    0x0014, 0x0015,  # Control board HW H+L
    0x0016, 0x0017,  # HMI HW H+L
    0x0018, 0x0019,  # Power board SW H+L
    0x001A, 0x001B,  # Control board SW H+L
    0x001C, 0x001D,  # HMI SW H+L
    0x001E, 0x001F,  # Version time H+L
    0x0020, 0x0021,  # CheckSum H+L
    0x0022,          # PC protocol version
    0x0023,          # WiFi protocol version
    0x0024,          # BMS protocol version
    0x0025,          # Parallel protocol version
    0x0038, 0x0039,  # Protocol ID A, B
    0x003A, 0x003B,  # Protocol ID C, D
    0x003C, 0x003D,  # Device type H+L
    0x003E, 0x003F,  # Company info H+L
    # WF_Simplify – Status
    0x0040,          # Software version
    0x0041,          # BMS connection status
    0x0042, 0x0043, 0x0044,
    0x0045,          # Parallel operation
    0x0046, 0x0047, 0x0048, 0x0049,
    # WF_Simplify – Grid Phase A
    0x0050, 0x0051, 0x0052, 0x0053,
    # WF_Simplify – Load Phase A
    0x0058, 0x0059, 0x005A, 0x005B, 0x005C, 0x005D,
    # WF_Simplify – Battery
    0x0080, 0x0081,
    0x0082, 0x0083,  # Negative battery voltage/current
    0x0084, 0x0085,
    # WF_Simplify – BMS
    0x0088, 0x0089, 0x008A, 0x008B,
    0x008C, 0x008D, 0x008E,  # BMS CV point, rated/current capacity
    0x008F,
    0x0090, 0x0091, 0x0092,  # BMS network status, failure, alarms
    # WF_Simplify – PV
    0x0096, 0x0097, 0x0098,
    0x0099, 0x009A, 0x009B,
    0x009C, 0x009D,
    # WF_LINE – Phase A mains (detailed)
    0x01B0, 0x01B1, 0x01B2, 0x01B3, 0x01B4,
    # WF_LINE – Grid energy counters (H+L pairs)
    0x01B7, 0x01B8,  # Feed-in today
    0x01B9, 0x01BA,  # Feed-in month
    0x01BB, 0x01BC,  # Feed-in year
    0x01BD, 0x01BE,  # Feed-in total
    0x01BF, 0x01C0,  # Consumption today
    0x01C1, 0x01C2,  # Consumption month
    0x01C3, 0x01C4,  # Consumption year
    0x01C5, 0x01C6,  # Consumption total
    # WF_OP – Output priority / mode
    0x0210, 0x0211,
    # WF_OP – Output Phase A
    0x0218, 0x0219, 0x021A, 0x021C, 0x021D, 0x0220,
    # WF_OP – Output energy counters (H+L pairs)
    0x0221, 0x0222,  # Output today
    0x0223, 0x0224,  # Output month
    0x0225, 0x0226,  # Output year
    0x0227, 0x0228,  # Output total
    # WF_PV – Total PV
    0x0260,          # Total PV voltage
    0x0261, 0x0262,  # Total PV current H+L
    0x0263, 0x0264,  # Total PV power H+L (existing only used H)
    0x0265, 0x0266,  # Total tracking status, rechargeable state
    0x0267, 0x0268,  # PV energy today H+L
    0x0269, 0x026A,  # PV energy month H+L
    0x026B, 0x026C,  # PV energy year H+L
    0x026D, 0x026E,  # PV energy total H+L
    # WF_PV – PV1 detailed
    0x0270, 0x0271, 0x0272,  # PV1 V, I, P (same data as 0x0096-0x0098)
    0x0273, 0x0274,  # PV1 tracking status, rechargeable state
    0x0275, 0x0276,  # PV1 energy today H+L
    0x0277, 0x0278,  # PV1 energy month H+L
    0x0279, 0x027A,  # PV1 energy year H+L
    0x027B, 0x027C,  # PV1 energy total H+L
    # WF_FAN
    0x0320, 0x0321,
    # WF_NTC – Temperatures
    0x0330, 0x0331, 0x0332, 0x0334,
    0x0333, 0x0335,  # CHG2 temp, DIS_CHG temp
}


# ── Sheet display order ────────────────────────────────────────────────────────
REGISTER_SHEETS = [
    ' WF_Version',
    ' WF_Simplify',
    ' WF_WorkMode',
    ' WF_BAT',
    ' WF_BMS',
    ' WF_LINE',
    ' WF_OP',
    ' WF_PV',
    ' WF_INV',
    ' WF_NTC',
    ' WF_FAN',
    ' WF_FunctionSwitch',
    ' WF_Time',
    ' WF_PARA',
    ' WF_Generator',
    ' WF_Admin',
    ' WF_FCT',
    ' WF_Calibration',
    ' WF_OTA',
    ' WF_UserSet',
]

LOGGER_RANGES = [
    (0x0000, 11), (0x0010, 48), (0x0040, 94), (0x0140, 5),
    (0x0150, 52), (0x0190, 19), (0x01B0, 83), (0x0210, 71),
    (0x0260, 34), (0x02A0, 54), (0x0320, 2), (0x0330, 5),
    (0x0350, 8), (0x0360, 6), (0x1000, 13), (0x1100, 10),
    (0x1200, 16), (0x2000, 3), (0x4100, 64), (0x4140, 12)
]
LOGGER_ADDRS = set()
for start, count in LOGGER_RANGES:
    for addr in range(start, start + count):
        LOGGER_ADDRS.add(addr)

# ── Regular (runtime) logger queries — parsed from uart_debug log ─────────────
REGULAR_RANGES = [
    (0x0039, 1), (0x0000, 10), (0x0010, 2), (0x001A, 2),
    (0x0043, 2), (0x0046, 4), (0x0050, 4), (0x0058, 6),
    (0x0080, 2), (0x0084, 2), (0x008F, 4), (0x0096, 6),
    (0x00A0, 1), (0x0140, 1), (0x0195, 3), (0x019B, 2),
    (0x01BF, 8), (0x0221, 8), (0x0267, 8),
]
REGULAR_ADDRS = set()
for start, count in REGULAR_RANGES:
    for addr in range(start, start + count):
        REGULAR_ADDRS.add(addr)

def safe(v, maxlen=130):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).replace('|', '/').replace('\n', ' ').strip()
    if len(s) > maxlen:
        s = s[:maxlen - 3] + '...'
    return s

lines = []

# ── Title ─────────────────────────────────────────────────────────────────────
lines.append('# JSD Solar 3KU — Повна карта Modbus регістрів')
lines.append('')
lines.append('> **Джерело:** `INV-Modbus地址表3KU（外发）V1.20.xlsx`  ')
lines.append('> **Тип:** Holding Registers, Function Code 03  ')
lines.append('> **Адреси:** 0-based hex (поле `address:` в ESPHome)  ')
lines.append('>')
lines.append('> **Колонка Initial Loger query:** позначає регістри, які логер запитує при запуску (за даними стартового лога)  ')
lines.append('> **Колонка Regular Loger query:** позначає регістри, які логер запитує у штатному циклі роботи (за даними uart_debug лога)  ')
lines.append('> **Колонка ✅:** позначає регістри, для яких вже створено сенсор у `jsd-solar-inverter.yaml`')
lines.append('')
lines.append('## Зміст')
lines.append('')
lines.append('1. [Версії документа](#version-information)')
lines.append('2. [Формат комунікації](#communication-format)')
lines.append('3. [Регістри](#register-sheets)')
lines.append('')

# ══════════════════════════════════════════════════════════════════════════════
# Section 1 – Version information
# ══════════════════════════════════════════════════════════════════════════════
lines.append('---')
lines.append('')
lines.append('## Version information')
lines.append('<a name="version-information"></a>')
lines.append('')
lines.append('| Версія | Автор | Дата | Зміни |')
lines.append('|--------|-------|------|-------|')

ws_ver = wb[' Version information']
for row in ws_ver.iter_rows(values_only=True):
    if row[0] is None or str(row[0]).lower() == 'version':
        continue
    ver   = safe(row[0], 10)
    auth  = safe(row[1], 40)
    date  = safe(row[2], 20)
    chng  = safe(row[3], 150)
    lines.append(f'| {ver} | {auth} | {date} | {chng} |')

lines.append('')

# ══════════════════════════════════════════════════════════════════════════════
# Section 2 – Communication format (вільний текст)
# ══════════════════════════════════════════════════════════════════════════════
lines.append('---')
lines.append('')
lines.append('## Communication format')
lines.append('<a name="communication-format"></a>')
lines.append('')
lines.append('**Параметри порту:** 9600 bps, 8N1 (8 data bits, No parity, 1 stop bit)')
lines.append('')

# Frame format table (rows 0-2)
lines.append('### Запит (Function Code 0x03 / 0x10)')
lines.append('')
lines.append('| Поле | Slave Addr | Func Code | Addr High | Addr Low | Reg Count High | Reg Count Low | CRC Low | CRC High |')
lines.append('|------|-----------|-----------|-----------|----------|---------------|--------------|---------|----------|')
lines.append('| Байти | 1 | 1 | 1 | 1 | 1 | 1 | 1 | 1 |')
lines.append('')

lines.append('### Доступні Function Codes')
lines.append('')
lines.append('| Код | Опис |')
lines.append('|-----|------|')
lines.append('| 0x03 | Читання Holding Registers (Read) |')
lines.append('| 0x10 | Запис Holding Registers (Write) |')
lines.append('')

lines.append('### Exception Codes')
lines.append('')
lines.append('| Код | Опис |')
lines.append('|-----|------|')
ws_comm = wb[' Communication format']
in_exception_block = False
for row in ws_comm.iter_rows(values_only=True):
    if row[0] is None:
        continue
    # Use raw string for keyword detection (safe() may truncate)
    raw0 = str(row[0]).lower()
    raw1 = str(row[1]).lower() if row[1] is not None else ''
    c0 = safe(row[0], 10)
    c1 = safe(row[1], 150)
    # Detect start of exception codes block
    if 'exception' in raw0 or 'exception' in raw1:
        in_exception_block = True
        continue
    # Stop at next section header (non-digit, non-empty)
    if in_exception_block and c0 and not c0.isdigit():
        in_exception_block = False
    # Collect numbered rows within the exception block
    if in_exception_block and c0.isdigit() and c1 and len(c1) > 10:
        lines.append(f'| {c0} | {c1} |')

lines.append('')

# ══════════════════════════════════════════════════════════════════════════════
# Section 3 – Register sheets
# ══════════════════════════════════════════════════════════════════════════════
lines.append('---')
lines.append('')
lines.append('## Регістри')
lines.append('<a name="register-sheets"></a>')
lines.append('')

total_regs   = 0
total_marked = 0

all_register_sheets = REGISTER_SHEETS + [
    s for s in wb.sheetnames
    if s not in REGISTER_SHEETS
    and s not in (' Version information', ' Communication format')
]

for sn in all_register_sheets:
    if sn not in wb.sheetnames:
        continue
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (must contain 'address in decimal')
    hdr_idx = None
    for i, row in enumerate(rows):
        if row and any(
            str(c).lower().strip() == 'address in decimal'
            for c in row if c is not None
        ):
            hdr_idx = i
            break
    if hdr_idx is None:
        continue

    # Collect data rows
    data_rows = []
    for row in rows[hdr_idx + 1:]:
        if row[1] is None:
            continue
        try:
            addr_dec = int(row[1])
        except (TypeError, ValueError):
            continue
        addr_hex_str = str(row[2]).upper().zfill(4) if row[2] else '????'
        try:
            addr_int = int(addr_hex_str, 16)
        except ValueError:
            addr_int = -1
        marked = (addr_int in SENSOR_ADDRS)
        in_logger_query = (addr_int in LOGGER_ADDRS)
        in_regular_query = (addr_int in REGULAR_ADDRS)
        data_rows.append({
            'dec':      addr_dec,
            'hex':      addr_hex_str,
            'marked':   marked,
            'in_logger_query': in_logger_query,
            'in_regular_query': in_regular_query,
            'meaning':  safe(row[5]),
            'type':     safe(row[6], 20),
            'unit':     safe(row[7], 20),
            'perm':     safe(row[4], 30),
            'usage':    safe(row[8], 130),
            'lo':       safe(row[9],  20),
            'hi':       safe(row[10], 20),
        })

    if not data_rows:
        continue

    sheet_label = sn.strip()
    lines.append('---')
    lines.append('')
    lines.append(f'### {sheet_label}')
    lines.append('')
    lines.append('| Initial Loger query | Regular Loger query | ✅ | Hex | Dec | Назва | Тип | Одиниця | Дозвіл | Примітки |')
    lines.append('|---|---|----|-----|-----|-------|-----|---------|--------|----------|')

    for r in data_rows:
        log_mark = '✅' if r['in_logger_query'] else ''
        reg_mark = '✅' if r['in_regular_query'] else ''
        mark = '✅' if r['marked'] else ''
        note = r['usage']
        if r['lo'] and r['hi']:
            note = (note + ' [' + r['lo'] + '…' + r['hi'] + ']').strip()
        row_str = (
            f"| {log_mark} "
            f"| {reg_mark} "
            f"| {mark} "
            f"| `0x{r['hex']}` "
            f"| {r['dec']} "
            f"| {r['meaning']} "
            f"| {r['type']} "
            f"| {r['unit']} "
            f"| {r['perm']} "
            f"| {note} |"
        )
        lines.append(row_str)
        total_regs += 1
        if r['marked']:
            total_marked += 1

    lines.append('')

lines.append('---')
lines.append('')
lines.append(
    f'*Всього регістрів: **{total_regs}** | '
    f'Реалізовано сенсорів у YAML: **{total_marked}***'
)

out_path = r'docs\registers-map.md'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'Done: {total_regs} registers, {total_marked} marked, {len(lines)} lines.')
